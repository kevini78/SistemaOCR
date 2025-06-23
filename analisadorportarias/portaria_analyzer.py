import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
from datetime import datetime, date
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
class PortariaAnalyzer:
    def __init__(self, planilha_historico_path=None):
        """
        Inicializa o analisador de portarias
        
        Args:
            planilha_historico_path (str): Caminho para a planilha com histórico de naturalizações (2018-2025)
        """
        self.planilha_historico_path = planilha_historico_path
        self.historico_df = None
        self.erros_encontrados = []
        self.paises_oficiais = self.carregar_paises_oficiais()
        
        # Carregar histórico se fornecido
        if planilha_historico_path and os.path.exists(planilha_historico_path):
            self.carregar_historico()
    
    def carregar_paises_oficiais(self):
        """Carrega lista de países oficiais"""
        return {
            'AFEGANISTÃO', 'ÁFRICA DO SUL', 'ALBÂNIA', 'ALEMANHA', 'ANDORRA', 'ANGOLA', 
            'ANTÍGUA E BARBUDA', 'ARÁBIA SAUDITA', 'ARGÉLIA', 'ARGENTINA', 'ARMÊNIA', 
            'AUSTRÁLIA', 'ÁUSTRIA', 'AZERBAIJÃO', 'BAHAMAS', 'BAHREIN', 'BANGLADESH', 
            'BARBADOS', 'BELARUS', 'BÉLGICA', 'BELIZE', 'BENIN', 'BOLÍVIA', 'BÓSNIA E HERZEGOVINA', 
            'BOTSUANA', 'BRASIL', 'BRUNEI', 'BULGÁRIA', 'BURKINA FASO', 'BURUNDI', 'BUTÃO', 
            'CABO VERDE', 'CAMARÕES', 'CAMBOJA', 'CANADÁ', 'CATAR', 'CAZAQUISTÃO', 'CHADE', 
            'CHILE', 'CHINA', 'CHIPRE', 'COLÔMBIA', 'COMORES', 'CONGO', 'COREIA DO NORTE', 
            'COREIA DO SUL', 'COSTA DO MARFIM', 'COSTA RICA', 'CROÁCIA', 'CUBA', 'DINAMARCA', 
            'DJIBUTI', 'DOMINICA', 'EGITO', 'EL SALVADOR', 'EMIRADOS ÁRABES UNIDOS', 'EQUADOR', 
            'ERITREIA', 'ESLOVÁQUIA', 'ESLOVÊNIA', 'ESPANHA', 'ESTADOS UNIDOS', 'ESTÔNIA', 
            'ESWATINI', 'ETIÓPIA', 'FIJI', 'FILIPINAS', 'FINLÂNDIA', 'FRANÇA', 'GABÃO', 
            'GÂMBIA', 'GANA', 'GEÓRGIA', 'GRANADA', 'GRÉCIA', 'GUATEMALA', 'GUIANA', 
            'GUINÉ', 'GUINÉ-BISSAU', 'GUINÉ EQUATORIAL', 'HAITI', 'HONDURAS', 'HUNGRIA', 
            'IÊMEN', 'ILHAS MARSHALL', 'ILHAS SALOMÃO', 'ÍNDIA', 'INDONÉSIA', 'IRÃ', 
            'IRAQUE', 'IRLANDA', 'ISLÂNDIA', 'ISRAEL', 'ITÁLIA', 'JAMAICA', 'JAPÃO', 
            'JORDÂNIA', 'KOSOVO', 'KUWAIT', 'LAOS', 'LESOTO', 'LETÔNIA', 'LÍBANO', 
            'LIBÉRIA', 'LÍBIA', 'LIECHTENSTEIN', 'LITUÂNIA', 'LUXEMBURGO', 'MACEDÔNIA DO NORTE', 
            'MADAGASCAR', 'MALÁSIA', 'MALAWI', 'MALDIVAS', 'MALI', 'MALTA', 'MARROCOS', 
            'MAURÍCIO', 'MAURITÂNIA', 'MÉXICO', 'MIANMAR', 'MICRONÉSIA', 'MOÇAMBIQUE', 
            'MOLDÁVIA', 'MÔNACO', 'MONGÓLIA', 'MONTENEGRO', 'NAMÍBIA', 'NAURU', 'NEPAL', 
            'NICARÁGUA', 'NÍGER', 'NIGÉRIA', 'NORUEGA', 'NOVA ZELÂNDIA', 'OMÃ', 'PAÍSES BAIXOS', 
            'PALAU', 'PALESTINA', 'PANAMÁ', 'PAPUA-NOVA GUINÉ', 'PAQUISTÃO', 'PARAGUAI', 'PERU', 'POLÔNIA', 
            'PORTUGAL', 'QUÊNIA', 'QUIRGUIZISTÃO', 'REINO UNIDO', 'REPÚBLICA CENTRO-AFRICANA', 
            'REPÚBLICA DEMOCRÁTICA DO CONGO', 'REPÚBLICA DOMINICANA', 'REPÚBLICA TCHECA', 
            'ROMÊNIA', 'RUANDA', 'RÚSSIA', 'SAMOA', 'SAN MARINO', 'SANTA LÚCIA', 
            'SÃO CRISTÓVÃO E NEVIS', 'SÃO TOMÉ E PRÍNCIPE', 'SÃO VICENTE E GRANADINAS', 
            'SEICHELES', 'SENEGAL', 'SERRA LEOA', 'SÉRVIA', 'SINGAPURA', 'SÍRIA', 
            'SOMÁLIA', 'SRI LANKA', 'SUAZILÂNDIA', 'SUDÃO', 'SUDÃO DO SUL', 'SUÉCIA', 
            'SUÍÇA', 'SURINAME', 'TAILÂNDIA', 'TAIWAN', 'TAJIQUISTÃO', 'TANZÂNIA', 
            'TIMOR-LESTE', 'TOGO', 'TONGA', 'TRINIDAD E TOBAGO', 'TUNÍSIA', 'TURCOMENISTÃO', 
            'TURQUIA', 'TUVALU', 'UCRÂNIA', 'UGANDA', 'URUGUAI', 'UZBEQUISTÃO', 'VANUATU', 
            'VATICANO', 'VENEZUELA', 'VIETNÃ', 'ZÂMBIA', 'ZIMBÁBUE'
        }
    
    def carregar_historico(self):
        """Carrega a planilha com histórico de naturalizações"""
        try:
            self.historico_df = pd.read_excel(self.planilha_historico_path)
            print(f"Histórico carregado: {len(self.historico_df)} registros")
        except Exception as e:
            print(f"Erro ao carregar histórico: {e}")
            self.historico_df = None
    
    def buscar_portaria_web(self, url_portaria):
        """
        Busca o conteúdo da portaria na web usando Selenium
        
        Args:
            url_portaria (str): URL da portaria
            
        Returns:
            str: Conteúdo HTML da portaria
        """
        print("Iniciando busca da portaria...")
        driver = None
        
        try:
            # Configurar Chrome
            print("Configurando Chrome WebDriver...")
            chrome_options = Options()
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
            
            # Tentar diferentes formas de inicializar o driver
            try:
                print("Tentando inicializar Chrome WebDriver...")
                driver = webdriver.Chrome(options=chrome_options)
                print("✅ Chrome WebDriver inicializado com sucesso!")
            except Exception as e1:
                print(f"❌ Erro ao inicializar Chrome: {e1}")
                
                # Tentar com service
                try:
                    print("Tentando com Service...")
                    service = Service()
                    driver = webdriver.Chrome(service=service, options=chrome_options)
                    print("✅ Chrome WebDriver inicializado com Service!")
                except Exception as e2:
                    print(f"❌ Erro com Service: {e2}")
                    
                    # Fallback para requests
                    print("Usando fallback com requests...")
                    return self.buscar_portaria_requests(url_portaria)
            
            # Acessar URL
            print(f"Acessando URL: {url_portaria}")
            driver.set_page_load_timeout(30)
            driver.get(url_portaria)
            
            print("Aguardando carregamento da página...")
            time.sleep(5)
            
            # Obter conteúdo
            print("Extraindo conteúdo...")
            content = driver.page_source
            
            print(f"✅ Conteúdo extraído! Tamanho: {len(content)} caracteres")
            
            return content
            
        except Exception as e:
            print(f"❌ Erro ao buscar portaria: {e}")
            print("Tentando método alternativo com requests...")
            return self.buscar_portaria_requests(url_portaria)
            
        finally:
            if driver:
                try:
                    driver.quit()
                    print("Chrome WebDriver fechado.")
                except:
                    pass
    
    def buscar_portaria_requests(self, url_portaria):
        """
        Método alternativo usando requests (fallback)
        """
        try:
            print("Usando método requests...")
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            response = requests.get(url_portaria, headers=headers, timeout=30)
            response.raise_for_status()
            
            print(f"✅ Conteúdo obtido via requests! Tamanho: {len(response.text)} caracteres")
            return response.text
            
        except Exception as e:
            print(f"❌ Erro com requests: {e}")
            return None
    
    def extrair_dados_portaria(self, html_content):
        """
        Extrai dados estruturados da portaria
        
        Args:
            html_content (str): Conteúdo HTML da portaria
            
        Returns:
            dict: Dados estruturados da portaria
        """
        print("Extraindo dados da portaria...")
        
        if not html_content:
            print("❌ Conteúdo HTML vazio")
            return None
        
        soup = BeautifulSoup(html_content, 'html.parser')
        texto_completo = soup.get_text()
        
        # Se não encontrou, tentar padrões alternativos
        if not match_portaria:
            padrao_alternativo = r'PORTARIA\s*(\d+[.,]?\d*)[\s,]*DE\s*(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})'
            match = re.search(padrao_alternativo, texto_portaria, re.IGNORECASE)
            if match:
                match_portaria = match
        
        if not match_portaria:
            print("❌ Nenhum padrão de portaria encontrado")
            return None
        
        numero_portaria = match_portaria.group(1)
        data_portaria = match_portaria.group(2)
        
        print(f"✅ Portaria encontrada: {numero_portaria} de {data_portaria}")
        
        # Identificar tipo de naturalização
        tipo_naturalizacao = self.identificar_tipo_naturalizacao(texto_portaria)
        print(f"✅ Tipo identificado: {tipo_naturalizacao}")
        
        # Extrair pessoas
        pessoas = self.extrair_pessoas(texto_portaria)
        print(f"✅ Pessoas extraídas: {len(pessoas)}")
        
        return {
            'numero': numero_portaria,
            'data': data_portaria,
            'tipo': tipo_naturalizacao,
            'pessoas': pessoas,
            'texto_completo': texto_portaria
        }
    
    def identificar_tipo_naturalizacao(self, texto):
        """Identifica o tipo de naturalização baseado no artigo e texto da portaria"""
        texto_lower = texto.lower()
        
        # Verificar se é tornar definitiva (art. 70 parágrafo único) - PRIORIDADE ALTA
        if 'tornar definitiva' in texto_lower and 'art. 70' in texto_lower:
            return 'DEFINITIVA'
        
        # Verificar se é naturalização provisória (art. 70) - PRIORIDADE ALTA
        elif 'art. 70' in texto_lower and 'naturalização provisória' in texto_lower:
            return 'PROVISORIA'
        
        # Verificar se é extraordinária (art. 67)
        elif 'art. 67' in texto_lower:
            return 'EXTRAORDINARIA'
        
        # Verificar se é ordinária (art. 65)
        elif 'art. 65' in texto_lower:
            return 'ORDINARIA'
        
        # Se não encontrou nenhum artigo específico, tentar identificar pelo contexto
        elif 'tornar definitiva' in texto_lower:
            return 'DEFINITIVA'
        elif 'naturalização provisória' in texto_lower:
            return 'PROVISORIA'
        elif 'extraordinária' in texto_lower:
            return 'EXTRAORDINARIA'
        elif 'ordinária' in texto_lower or 'por naturalização' in texto_lower:
            return 'ORDINARIA'
        
        return 'DESCONHECIDO'
    
    def extrair_pessoas(self, texto):
        """Extrai dados das pessoas da portaria"""
        print("Extraindo dados das pessoas...")
        pessoas = []
        
        # Padrões para extrair informações - mais flexíveis
        patterns = [
            # Padrão principal: NOME - DOCUMENTO, natural do PAÍS, nascido em DATA, filho de PAI
            r'([A-ZÀ-Ú][A-ZÀ-Ú\s]+?)\s*[-–,]\s*([A-Z]\d+[-]\w+|Processo\s+[][\d\.\/]+)\s*,?\s*natural\s+d[oa]\s+([A-ZÀ-Ú\s]+?)\s*,?\s*nascid[oa]\s+em\s+(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})\s*,?\s*filh[oa]\s+de\s+([A-ZÀ-Ú\s]+)',
            
            # Padrão alternativo sem "filho de"
            r'([A-ZÀ-Ú][A-ZÀ-Ú\s]+?)\s*[-–,]\s*([A-Z]\d+[-]\w+|Processo\s+[][\d\.\/]+)\s*,?\s*natural\s+d[oa]\s+([A-ZÀ-Ú\s]+?)\s*,?\s*nascid[oa]\s+em\s+(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})',
            
            # Padrão para casos onde o documento vem antes do nome
            r'([A-Z]\d+[-]\w+|Processo\s+[][\d\.\/]+)\s*,?\s*([A-ZÀ-Ú][A-ZÀ-Ú\s]+?)\s*,?\s*natural\s+d[oa]\s+([A-ZÀ-Ú\s]+)',
            
            # Padrão mais simples
            r'([A-ZÀ-Ú][A-ZÀ-Ú\s]+?)\s*[-–,]\s*([A-Z]\d+[-]\w+)\s*,?\s*natural\s+d[oa]\s+([A-ZÀ-Ú\s]+)',
            
            # Padrão para casos onde o nome vem depois do documento
            r'([A-Z]\d+[-]\w+)\s*,?\s*([A-ZÀ-Ú][A-ZÀ-Ú\s]+?)\s*,?\s*natural\s+d[oa]\s+([A-ZÀ-Ú\s]+)',
            
            # Padrão para casos onde não há hífen e nem vírgula
            r'([A-ZÀ-Ú][A-ZÀ-Ú\s]+?)\s+([A-Z]\d+[-]\w+|Processo\s+[][\d\.\/]+)\s+natural\s+d[oa]\s+([A-ZÀ-Ú\s]+)',
            
            # Padrão para casos onde o documento vem antes do nome (sem vírgulas)
            r'([A-Z]\d+[-]\w+|Processo\s+[][\d\.\/]+)\s+([A-ZÀ-Ú][A-ZÀ-Ú\s]+?)\s+natural\s+d[oa]\s+([A-ZÀ-Ú\s]+)',
            
            # Padrão alternativo para nomes com apóstrofo
            r'([A-ZÀ-Ú][A-ZÀ-Ú\s\']+?)\s*[-–,]\s*([A-Z]\d+[-]\w+|Processo\s+[][\d\.\/]+)\s*,?\s*natural\s+d[oa]\s+([A-ZÀ-Ú\s\']+?)\s*,?\s*nascid[oa]\s+em\s+(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})\s*,?\s*filh[oa]\s+de\s+([A-ZÀ-Ú\s\']+)',
            
            # Padrão para nomes com hífen
            r'([A-ZÀ-Ú][A-ZÀ-Ú\s-]+?)\s*[-–,]\s*([A-Z]\d+[-]\w+|Processo\s+[][\d\.\/]+)\s*,?\s*natural\s+d[oa]\s+([A-ZÀ-Ú\s-]+?)\s*,?\s*nascid[oa]\s+em\s+(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})\s*,?\s*filh[oa]\s+de\s+([A-ZÀ-Ú\s-]+)',
        ]

        # Tentar cada padrão
        for pattern in patterns:
            print(f"Tentando padrão: {pattern}")
            matches = re.findall(pattern, texto, re.IGNORECASE)
            
            if matches:
                print(f"✅ Encontradas {len(matches)} pessoas com padrão")
                
                for match in matches:
                    nome = match[0].strip()
                    documento = match[1].strip()
                    pais = match[2].strip().upper()
                    
                    # Verificar se temos data de nascimento
                    data_nascimento = match[3].strip() if len(match) > 3 else None
                    nome_pai = match[4].strip() if len(match) > 4 else None
                    
                    # Limpar nome (remover "e" no final se for apenas isso)
                    if nome.strip() == "e" or nome.strip() == "E":
                        continue
                    
                    # Verificar se é processo ou RNM
                    tipo_documento = 'PROCESSO' if 'Processo' in documento else 'RNM'
                    
                    # Calcular idade se temos data
                    idade = self.calcular_idade(data_nascimento) if data_nascimento else None
                    
                    pessoa = {
                        'nome': nome,
                        'documento': documento,
                        'tipo_documento': tipo_documento,
                        'pais': pais,
                        'data_nascimento': data_nascimento,
                        'nome_pai': nome_pai,
                        'idade': idade
                    }
                    
                    pessoas.append(pessoa)
                    print(f"  - {nome} ({pais})")
                
                break  # Se encontrou com um padrão, não tenta os outros
        
        # Limpar pessoas duplicadas ou com nomes inválidos
        pessoas_limpas = []
        nomes_vistos = set()
        
        for pessoa in pessoas:
            nome_limpo = pessoa['nome'].strip()
            # Remover "e" isolado, nomes muito curtos e nomes que começam com "e"
            if (nome_limpo and 
                nome_limpo not in nomes_vistos and 
                nome_limpo not in ['e', 'E'] and
                len(nome_limpo) > 2 and
                not nome_limpo.startswith('e ') and
                not nome_limpo.startswith('E ')):
                nomes_vistos.add(nome_limpo)
                pessoas_limpas.append(pessoa)
        
        if not pessoas_limpas:
            print("❌ Nenhuma pessoa encontrada. Tentando busca manual...")
            # Busca manual por nomes em maiúscula seguidos de hífen ou vírgula
            linhas = texto.split('\n')
            for linha in linhas:
                if re.search(r'[A-ZÀ-Ú\s]+[-–,]\s*[A-Z]\d+', linha):
                    print(f"Linha suspeita: {linha[:100]}...")
            
            # Tentar padrão mais simples: apenas nomes em maiúscula
            print("Tentando padrão simples de nomes...")
            nomes_simples = re.findall(r'([A-ZÀ-Ú][A-ZÀ-Ú\s]{2,})', texto)
            nomes_filtrados = [nome.strip() for nome in nomes_simples if len(nome.strip()) > 3]
            print(f"Nomes encontrados (simples): {nomes_filtrados[:10]}...")
        
        print(f"Total de pessoas extraídas: {len(pessoas_limpas)}")
        return pessoas_limpas
    
    def calcular_idade(self, data_nascimento_str):
        """Calcula idade baseada na data de nascimento em formato brasileiro"""
        try:
            # Converter data brasileira para datetime
            meses = {
                'janeiro': 1, 'fevereiro': 2, 'março': 3, 'abril': 4,
                'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
                'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12
            }
            
            partes = data_nascimento_str.lower().replace(' de ', ' ').split()
            dia = int(partes[0])
            mes = meses.get(partes[1], 1)
            ano = int(partes[2])
            
            data_nascimento = date(ano, mes, dia)
            hoje = date.today()
            
            idade = hoje.year - data_nascimento.year
            if hoje.month < data_nascimento.month or (hoje.month == data_nascimento.month and hoje.day < data_nascimento.day):
                idade -= 1
                
            return idade
        except:
            return None
    
    def verificar_erros(self, dados_portaria):
        """Verifica erros na portaria baseado nas regras"""
        erros = []
        
        if not dados_portaria:
            return [{'tipo': 'ERRO_PARSING', 'descrição': 'Não foi possível extrair dados da portaria'}]
        
        tipo = dados_portaria['tipo']
        pessoas = dados_portaria['pessoas']
        
        print(f"Verificando {len(pessoas)} pessoas para portaria tipo: {tipo}")
        
        # Verificar duplicatas na mesma portaria
        nomes_vistos = {}
        for pessoa in pessoas:
            nome_data = f"{pessoa['nome']}|{pessoa['data_nascimento']}"
            if nome_data in nomes_vistos:
                erros.append({
                    'tipo': 'DUPLICATA_MESMA_PORTARIA',
                    'pessoa': pessoa['nome'],
                    'descrição': f'{pessoa["nome"]} aparece mais de uma vez na mesma portaria'
                })
            nomes_vistos[nome_data] = True
        
        # Verificar cada pessoa
        for pessoa in pessoas:
            print(f"Verificando pessoa: {pessoa['nome']} (idade: {pessoa['idade']})")
            
            # Verificar idade conforme tipo de naturalização
            if tipo == 'ORDINARIA' and pessoa['idade'] and pessoa['idade'] < 18:
                erros.append({
                    'tipo': 'IDADE_INCORRETA_ORDINARIA',
                    'pessoa': pessoa['nome'],
                    'idade': pessoa['idade'],
                    'descrição': f'Naturalização ordinária (art. 65) com pessoa menor de 18 anos: {pessoa["nome"]} ({pessoa["idade"]} anos)'
                })
            
            elif tipo == 'EXTRAORDINARIA' and pessoa['idade'] and pessoa['idade'] < 15:
                erros.append({
                    'tipo': 'IDADE_INCORRETA_EXTRAORDINARIA',
                    'pessoa': pessoa['nome'],
                    'idade': pessoa['idade'],
                    'descrição': f'Naturalização extraordinária (art. 67) com pessoa menor de 15 anos: {pessoa["nome"]} ({pessoa["idade"]} anos)'
                })
            
            elif tipo == 'PROVISORIA' and pessoa['idade'] and pessoa['idade'] >= 18:
                erros.append({
                    'tipo': 'IDADE_INCORRETA_PROVISORIA',
                    'pessoa': pessoa['nome'],
                    'idade': pessoa['idade'],
                    'descrição': f'Naturalização provisória (art. 70) com pessoa maior ou igual a 18 anos: {pessoa["nome"]} ({pessoa["idade"]} anos)'
                })
            
            elif tipo == 'DEFINITIVA' and pessoa['idade'] and (pessoa['idade'] < 18 or pessoa['idade'] > 20):
                erros.append({
                    'tipo': 'IDADE_INCORRETA_DEFINITIVA',
                    'pessoa': pessoa['nome'],
                    'idade': pessoa['idade'],
                    'descrição': f'Naturalização definitiva com idade fora do intervalo 18-20 anos: {pessoa["nome"]} ({pessoa["idade"]} anos)'
                })
            
            # Verificar documento (Processo ou RNM)
            if pessoa['tipo_documento'] not in ['PROCESSO', 'RNM']:
                erros.append({
                    'tipo': 'DOCUMENTO_INVALIDO',
                    'pessoa': pessoa['nome'],
                    'descrição': f'Documento inválido para {pessoa["nome"]}: {pessoa["documento"]}'
                })
            
            # Verificar país
            if pessoa['pais'] not in self.paises_oficiais:
                erros.append({
                    'tipo': 'PAIS_INVALIDO',
                    'pessoa': pessoa['nome'],
                    'pais': pessoa['pais'],
                    'descrição': f'País não reconhecido: {pessoa["pais"]} para {pessoa["nome"]}'
                })
            
            # Verificar histórico de naturalizações anteriores
            if self.historico_df is not None:
                print(f"\n🔍 Verificando histórico para {pessoa['nome']}...")
                resultado_historico = self.verificar_duplicata_historico(pessoa)
                print(f"   Resultado: {resultado_historico}")
                
                if resultado_historico == "HISTORICO_NAO_CARREGADO":
                    erros.append({
                        'tipo': 'HISTORICO_NAO_DISPONIVEL',
                        'pessoa': pessoa['nome'],
                        'descrição': f'Histórico de naturalizações não disponível para verificar {pessoa["nome"]}'
                    })
                elif resultado_historico == "COLUNAS_INEXISTENTES":
                    erros.append({
                        'tipo': 'HISTORICO_ESTRUTURA_INVALIDA',
                        'pessoa': pessoa['nome'],
                        'descrição': f'Estrutura da planilha de histórico inválida para verificar {pessoa["nome"]}'
                    })
                elif resultado_historico != "NAO_PUBLICADO_ANTERIORMENTE":
                    erros.append({
                        'tipo': 'JA_NATURALIZADO_ANTERIORMENTE',
                        'pessoa': pessoa['nome'],
                        'descrição': f'{pessoa["nome"]} já foi naturalizado anteriormente em {resultado_historico}'
                    })
                else:
                    print(f"   ✅ {pessoa['nome']} não encontrado no histórico")
                # Se não foi publicado anteriormente, não adiciona nada
            else:
                print(f"⚠️  Histórico não disponível para verificar {pessoa['nome']}")
        
        return erros
    
    def verificar_duplicata_historico(self, pessoa):
        """Verifica se a pessoa já foi naturalizada anteriormente - VERIFICAÇÃO RIGOROSA"""
        if self.historico_df is None:
            print(f"❌ Histórico não carregado para verificar {pessoa['nome']}")
            return "HISTORICO_NAO_CARREGADO"
        
        print(f"🔍 Verificando histórico para: {pessoa['nome']} (Data: {pessoa['data_nascimento']})")
        
        # Verificar se as colunas existem
        colunas_necessarias = ['NOME COMPLETO', 'DATA DE NASCIMENTO']
        colunas_faltando = [col for col in colunas_necessarias if col not in self.historico_df.columns]
        
        if colunas_faltando:
            print(f"❌ Colunas faltando no histórico: {colunas_faltando}")
            print(f"Colunas disponíveis: {list(self.historico_df.columns)}")
            return "COLUNAS_INEXISTENTES"
        
        # Limpar e normalizar o nome para busca
        nome_busca = pessoa['nome'].upper().strip()
        data_busca = pessoa['data_nascimento']
        
        print(f"   Nome para busca: '{nome_busca}'")
        print(f"   Data para busca: '{data_busca}'")
        
        # Se não tem data de nascimento, NÃO verificar duplicata
        if not data_busca:
            print(f"   ⚠️  Sem data de nascimento, não é possível verificar duplicata com segurança")
            return "NAO_PUBLICADO_ANTERIORMENTE"
        
        # Normalizar formato da data para comparação
        data_normalizada = self.normalizar_data(data_busca)
        print(f"   Data normalizada: '{data_normalizada}'")
        
        # BUSCA RIGOROSA: apenas nome EXATO E data EXATA
        mask_exato = (
            (self.historico_df['NOME COMPLETO'].str.upper() == nome_busca) &
            (self.historico_df['DATA DE NASCIMENTO'].astype(str).apply(self.normalizar_data_planilha) == data_normalizada)
        )
        resultados_exatos = self.historico_df[mask_exato]
        
        print(f"   Encontrados {len(resultados_exatos)} registros com nome E data EXATOS")
        
        if len(resultados_exatos) > 0:
            primeira_duplicata = resultados_exatos.iloc[0]
            portaria = primeira_duplicata.get('Nº da Portaria', 'N/A')
            mes = primeira_duplicata.get('Mês', 'N/A')
            ano = primeira_duplicata.get('Ano', 'N/A')
            
            resultado = f"Portaria {portaria} ({mes}/{ano})"
            print(f"   ✅ DUPLICATA CONFIRMADA: {resultado}")
            return resultado
        
        # Se não encontrou, verificar se há nome exato mas data diferente (para debug)
        mask_nome_exato = self.historico_df['NOME COMPLETO'].str.upper() == nome_busca
        resultados_nome_exato = self.historico_df[mask_nome_exato]
        
        if len(resultados_nome_exato) > 0:
            primeira_duplicata = resultados_nome_exato.iloc[0]
            data_historico = primeira_duplicata.get('DATA DE NASCIMENTO', 'N/A')
            print(f"   ⚠️  Nome encontrado mas data diferente: '{data_historico}' vs '{data_normalizada}'")
        
        print(f"   ✅ Nenhuma duplicata encontrada")
        return "NAO_PUBLICADO_ANTERIORMENTE"
    
    def normalizar_data(self, data_str):
        """Normaliza o formato da data para comparação"""
        if not data_str:
            return ""
        
        # Converter para string e limpar
        data_str = str(data_str).strip()
        
        # Se já está no formato correto, retornar
        if re.match(r'\d{1,2}\s+de\s+\w+\s+de\s+\d{4}', data_str, re.IGNORECASE):
            return data_str
        
        # Tentar converter outros formatos
        try:
            # Converter data brasileira para datetime e de volta para formato padrão
            meses = {
                'janeiro': 1, 'fevereiro': 2, 'março': 3, 'abril': 4,
                'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
                'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12
            }
            
            # Verificar se é formato DD/MM/YYYY ou DD/M/YYYY
            if re.match(r'\d{1,2}/\d{1,2}/\d{4}', data_str):
                partes = data_str.split('/')
                dia = int(partes[0])
                mes = int(partes[1])
                ano = int(partes[2])
                
                # Retornar no formato padrão
                meses_nomes = ['janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
                              'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
                mes_nome = meses_nomes[mes - 1]
                
                return f"{dia} de {mes_nome} de {ano}"
            
            # Verificar se é formato DD de MES de YYYY
            partes = data_str.lower().replace(' de ', ' ').split()
            if len(partes) >= 3:
                dia = int(partes[0])
                mes = meses.get(partes[1], 1)
                ano = int(partes[2])
                
                # Retornar no formato padrão
                meses_nomes = ['janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
                              'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
                mes_nome = meses_nomes[mes - 1]
                
                return f"{dia} de {mes_nome} de {ano}"
        except:
            pass
        
        return data_str
    
    def normalizar_data_planilha(self, data_str):
        """Normaliza o formato da data da planilha para comparação"""
        if not data_str:
            return ""
        
        # Converter para string e limpar
        data_str = str(data_str).strip()
        
        # Se já está no formato correto, retornar
        if re.match(r'\d{1,2}\s+de\s+\w+\s+de\s+\d{4}', data_str, re.IGNORECASE):
            return data_str
        
        # Verificar se é formato DD/MM/YYYY ou DD/M/YYYY
        if re.match(r'\d{1,2}/\d{1,2}/\d{4}', data_str):
            partes = data_str.split('/')
            dia = int(partes[0])
            mes = int(partes[1])
            ano = int(partes[2])
            
            # Retornar no formato padrão
            meses_nomes = ['janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
                          'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
            mes_nome = meses_nomes[mes - 1]
            
            return f"{dia} de {mes_nome} de {ano}"
        
        # Verificar se é datetime
        if '00:00:00' in data_str:
            try:
                from datetime import datetime
                dt = datetime.strptime(data_str, '%Y-%m-%d %H:%M:%S')
                meses_nomes = ['janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
                              'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
                mes_nome = meses_nomes[dt.month - 1]
                return f"{dt.day} de {mes_nome} de {dt.year}"
            except:
                pass
        
        return data_str
    
    def nomes_similares(self, nome1, nome2):
        """Verifica se dois nomes são similares"""
        nome1_clean = re.sub(r'\s+', ' ', nome1.upper().strip())
        nome2_clean = re.sub(r'\s+', ' ', nome2.upper().strip())
        
        # Verificar se são iguais
        if nome1_clean == nome2_clean:
            return True
        
        # Verificar se um contém o outro
        if nome1_clean in nome2_clean or nome2_clean in nome1_clean:
            return True
        
        # Verificar similaridade por palavras (pelo menos 2 palavras iguais)
        palavras1 = set(nome1_clean.split())
        palavras2 = set(nome2_clean.split())
        
        # Remover palavras muito comuns que podem causar falsos positivos
        palavras_comuns = {'DA', 'DE', 'DO', 'DAS', 'DOS', 'E'}
        palavras1 = palavras1 - palavras_comuns
        palavras2 = palavras2 - palavras_comuns
        
        palavras_comuns = palavras1.intersection(palavras2)
        
        # Para nomes com muitas palavras, exigir pelo menos 3 palavras iguais
        if len(palavras1) >= 4 or len(palavras2) >= 4:
            if len(palavras_comuns) >= 3:
                return True
        # Para nomes menores, 2 palavras iguais são suficientes
        elif len(palavras_comuns) >= 2:
            return True
        
        # Verificar se há pelo menos 70% de similaridade de caracteres
        if len(nome1_clean) > 10 and len(nome2_clean) > 10:
            from difflib import SequenceMatcher
            similaridade = SequenceMatcher(None, nome1_clean, nome2_clean).ratio()
            if similaridade >= 0.7:
                return True
        
        return False
    
    def gerar_relatorio_excel(self, dados_portaria, erros, nome_arquivo=None):
        """Gera relatório em Excel com os erros encontrados"""
        if nome_arquivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"relatorio_erros_portaria_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Aba 1: Resumo
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        
        # Cabeçalho do resumo
        ws_resumo['A1'] = "RELATÓRIO DE ANÁLISE DE PORTARIA"
        ws_resumo['A1'].font = Font(bold=True, size=14)
        
        ws_resumo['A3'] = "Número da Portaria:"
        ws_resumo['B3'] = dados_portaria['numero'] if dados_portaria else "N/A"
        
        ws_resumo['A4'] = "Data:"
        ws_resumo['B4'] = dados_portaria['data'] if dados_portaria else "N/A"
        
        ws_resumo['A5'] = "Tipo de Naturalização:"
        ws_resumo['B5'] = dados_portaria['tipo'] if dados_portaria else "N/A"
        
        ws_resumo['A6'] = "Total de Pessoas:"
        ws_resumo['B6'] = len(dados_portaria['pessoas']) if dados_portaria else 0
        
        ws_resumo['A7'] = "Total de Erros:"
        ws_resumo['B7'] = len(erros)
        
        # Aba 2: Erros Detalhados
        ws_erros = wb.create_sheet(title="Erros")
        
        # Cabeçalhos
        headers_erros = ['Tipo de Erro', 'Pessoa', 'Idade', 'País', 'Descrição']
        for col, header in enumerate(headers_erros, 1):
            cell = ws_erros.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Dados dos erros
        for row, erro in enumerate(erros, 2):
            ws_erros.cell(row=row, column=1, value=erro['tipo'])
            ws_erros.cell(row=row, column=2, value=erro.get('pessoa', ''))
            ws_erros.cell(row=row, column=3, value=erro.get('idade', ''))
            ws_erros.cell(row=row, column=4, value=erro.get('pais', ''))
            ws_erros.cell(row=row, column=5, value=erro['descrição'])
        
        # Aba 3: Pessoas da Portaria
        if dados_portaria and dados_portaria['pessoas']:
            ws_pessoas = wb.create_sheet(title="Pessoas")
            
            headers_pessoas = ['Nome', 'Documento', 'Tipo Doc', 'País', 'Data Nascimento', 'Idade', 'Nome do Pai']
            for col, header in enumerate(headers_pessoas, 1):
                cell = ws_pessoas.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            for row, pessoa in enumerate(dados_portaria['pessoas'], 2):
                ws_pessoas.cell(row=row, column=1, value=pessoa['nome'])
                ws_pessoas.cell(row=row, column=2, value=pessoa['documento'])
                ws_pessoas.cell(row=row, column=3, value=pessoa['tipo_documento'])
                ws_pessoas.cell(row=row, column=4, value=pessoa['pais'])
                ws_pessoas.cell(row=row, column=5, value=pessoa['data_nascimento'])
                ws_pessoas.cell(row=row, column=6, value=pessoa['idade'])
                ws_pessoas.cell(row=row, column=7, value=pessoa['nome_pai'])
        
        # Salvar arquivo
        wb.save(nome_arquivo)
        print(f"Relatório salvo em: {nome_arquivo}")
        
        return nome_arquivo
    
    def analisar_portaria(self, url_portaria, gerar_excel=True):
        """
        Método principal para analisar uma portaria
        
        Args:
            url_portaria (str): URL da portaria
            gerar_excel (bool): Se deve gerar relatório em Excel
            
        Returns:
            dict: Resultado da análise
        """
        print(f"Analisando portaria: {url_portaria}")
        
        # 1. Buscar conteúdo da portaria
        html_content = self.buscar_portaria_web(url_portaria)
        if not html_content:
            return {'erro': 'Não foi possível acessar a portaria'}
        
        # 2. Extrair texto do HTML
        soup = BeautifulSoup(html_content, 'html.parser')
        texto_completo = soup.get_text()
        
        # 3. Analisar múltiplas portarias se houver
        resultados, arquivos_excel = self.analisar_multiplas_portarias(texto_completo, gerar_excel)
        
        # 4. Retornar resultado consolidado
        total_erros = sum(r['total_erros'] for r in resultados)
        
        return {
            'resultados': resultados,
            'arquivos_excel': arquivos_excel,
            'total_erros': total_erros,
            'total_portarias': len(resultados)
        }
    
    def analisar_texto_portaria(self, texto_portaria, gerar_excel=True):
        """
        Método para analisar uma portaria a partir do texto direto
        
        Args:
            texto_portaria (str): Texto da portaria
            gerar_excel (bool): Se deve gerar relatório em Excel
            
        Returns:
            dict: Resultado da análise
        """
        print(f"Analisando texto da portaria...")
        
        # 1. Extrair dados estruturados
        dados_portaria = self.extrair_dados_portaria_direto(texto_portaria)
        
        # 2. Verificar erros
        erros = self.verificar_erros(dados_portaria)
        
        # 3. Gerar relatório
        if gerar_excel:
            arquivo_excel = self.gerar_relatorio_excel(dados_portaria, erros)
        else:
            arquivo_excel = None
        
        # 4. Mostrar resumo
        print("\n" + "="*50)
        print("RESUMO DA ANÁLISE")
        print("="*50)
        
        if dados_portaria:
            print(f"Portaria: {dados_portaria['numero']}")
            print(f"Data: {dados_portaria['data']}")
            print(f"Tipo: {dados_portaria['tipo']}")
            print(f"Total de pessoas: {len(dados_portaria['pessoas'])}")
        
        print(f"Total de erros encontrados: {len(erros)}")
        
        if erros:
            print("\nERROS ENCONTRADOS:")
            for i, erro in enumerate(erros, 1):
                print(f"{i}. {erro['descrição']}")
        else:
            print("\n✅ Nenhum erro encontrado!")
        
        return {
            'dados_portaria': dados_portaria,
            'erros': erros,
            'arquivo_excel': arquivo_excel,
            'total_erros': len(erros)
        }
    
    def formatar_numero_portaria(self, numero):
        """
        Formata o número da portaria mantendo o formato original
        
        Args:
            numero (str): Número da portaria
            
        Returns:
            str: Número formatado
        """
        # Remover pontos e vírgulas temporariamente para converter para número
        numero_limpo = numero.replace('.', '').replace(',', '')
        numero_int = int(float(numero_limpo))
        
        # Formatar com pontos e vírgulas conforme o padrão oficial
        numero_formatado = format(numero_int, ',').replace(',', '.')
        return f"Nº {numero_formatado}".upper()

    def extrair_dados_portaria_direto(self, texto_portaria):
        """
        Extrai dados estruturados da portaria a partir do texto direto
        
        Args:
            texto_portaria (str): Texto da portaria
            
        Returns:
            dict: Dados estruturados da portaria
        """
        print("Extraindo dados da portaria...")
        
        if not texto_portaria:
            print("❌ Texto da portaria vazio")
            return None
        
        # Primeiro, encontrar o número e data da portaria
        match_portaria = None
        
        # Tentar encontrar o padrão principal
        padrao_principal = r'PORTARIA\s*N[º°]?\s*(\d+[.,]?\d*)[\s,]*DE\s*(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})'
        match = re.search(padrao_principal, texto_portaria, re.IGNORECASE)
        if match:
            match_portaria = match
        
        # Se não encontrou, tentar padrões alternativos
        if not match_portaria:
            padrao_alternativo = r'PORTARIA\s*(\d+[.,]?\d*)[\s,]*DE\s*(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})'
            match = re.search(padrao_alternativo, texto_portaria, re.IGNORECASE)
            if match:
                match_portaria = match
        
        if not match_portaria:
            print("❌ Nenhum padrão de portaria encontrado")
            return None
        
        numero_portaria = match_portaria.group(1)
        data_portaria = match_portaria.group(2)
        
        print(f"✅ Portaria encontrada: {numero_portaria} de {data_portaria}")
        
        # Identificar tipo de naturalização
        tipo_naturalizacao = self.identificar_tipo_naturalizacao(texto_portaria)
        print(f"✅ Tipo identificado: {tipo_naturalizacao}")
        
        # Extrair pessoas
        pessoas = self.extrair_pessoas(texto_portaria)
        print(f"✅ Pessoas extraídas: {len(pessoas)}")
        
        return {
            'numero': numero_portaria,
            'data': data_portaria,
            'tipo': tipo_naturalizacao,
            'pessoas': pessoas,
            'texto_completo': texto_portaria
        }

    def analisar_multiplas_portarias(self, texto_completo, gerar_excel=True):
        """
        Analisa múltiplas portarias em um texto único, separando cada portaria e identificando o tipo corretamente.
        """
        # Regex para separar portarias de forma mais precisa
        # Captura desde PORTARIA até o início de uma nova portaria ou fim do texto
        # Adiciona verificação para não capturar referências à portaria 623
        padrao_portaria = r'(PORTARIA\s*,?\s*Nº\s*(\d+[.,]?\d*)\s*,?\s*DE\s*(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})[\s\S]*?(?=PORTARIA\s*,?\s*Nº\s*(\d+[.,]?\d*)\s*,?\s*DE\s*(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})|$))'
        
        # Encontrar todas as posições onde começam portarias
        matches = list(re.finditer(padrao_portaria, texto_completo))
        
        if not matches:
            print("Nenhuma portaria encontrada no texto")
            print("Debug: Primeiros 500 caracteres do texto:")
            print(texto_completo[:500])
            return [], []
        
        print(f"Encontradas {len(matches)} portarias no texto")
        
        # Extrair os blocos completos usando os matches
        blocos = [match.group(1) for match in matches]  # Usar group(1) para pegar o bloco completo
        
        # Lista para rastrear todas as pessoas do documento
        todas_pessoas_documento = []
        resultados = []
        arquivos_excel = []
        
        for i, bloco in enumerate(blocos, 1):
            print(f"\n==============================")
            print(f"Analisando PORTARIA {i}...")
            print(f"Tamanho do bloco: {len(bloco)} caracteres")
            print(f"Primeiros 200 caracteres: {bloco[:200]}...")
            
            dados_portaria = self.extrair_dados_portaria_direto(bloco)
            if not dados_portaria:
                print(f"❌ Não foi possível extrair dados da portaria {i}")
                continue
                
            # Verificar se a portaria tem pessoas antes de processar
            if not dados_portaria.get('pessoas') or len(dados_portaria['pessoas']) == 0:
                print(f"⚠️  Portaria {i} não tem pessoas, pulando...")
                continue
                
            erros = self.verificar_erros(dados_portaria)
            
            # Verificar duplicatas entre portarias do mesmo documento
            erros_duplicatas = self.verificar_duplicatas_entre_portarias(dados_portaria, todas_pessoas_documento)
            erros.extend(erros_duplicatas)
            
            # Adicionar pessoas desta portaria à lista geral
            for pessoa in dados_portaria['pessoas']:
                todas_pessoas_documento.append({
                    'nome': pessoa['nome'],
                    'data_nascimento': pessoa['data_nascimento'],
                    'portaria_origem': dados_portaria['numero']
                })
            
            arquivo_excel = None
            if gerar_excel:
                arquivo_excel = self.gerar_relatorio_excel(dados_portaria, erros)
                arquivos_excel.append(arquivo_excel)
            
            resultados.append({
                'dados_portaria': dados_portaria,
                'erros': erros,
                'arquivo_excel': arquivo_excel,
                'total_erros': len(erros)
            })
        
        return resultados, arquivos_excel

    def verificar_duplicatas_entre_portarias(self, dados_portaria_atual, todas_pessoas_documento):
        """
        Verifica se pessoas da portaria atual já apareceram em outras portarias do mesmo documento
        """
        erros = []
        
        for pessoa_atual in dados_portaria_atual['pessoas']:
            nome_atual = pessoa_atual['nome']
            data_atual = pessoa_atual['data_nascimento']
            
            # Verificar se esta pessoa já apareceu em outras portarias do documento
            for pessoa_anterior in todas_pessoas_documento:
                if (pessoa_anterior['nome'] == nome_atual and 
                    pessoa_anterior['data_nascimento'] == data_atual):
                    
                    erros.append({
                        'tipo': 'DUPLICATA_ENTRE_PORTARIAS',
                        'pessoa': nome_atual,
                        'descrição': f'{nome_atual} aparece em mais de uma portaria no mesmo documento: Portaria {pessoa_anterior["portaria_origem"]} e Portaria {dados_portaria_atual["numero"]}'
                    })
                    break  # Só precisa encontrar uma duplicata
        
        return erros

# Exemplo de uso - versão melhorada com diagnósticos
def main():
    print("="*60)
    print("🔍 ANALISADOR DE PORTARIAS DE NATURALIZAÇÃO")
    print("="*60)
    
    # Configurar analisador
    print("\n1. Configurando analisador...")
    
    # Verificar se há planilha de histórico
    arquivos_historico = [
        "historico_naturalizacoes.xlsx"
    ]
    
    caminho_historico = None
    for arquivo in arquivos_historico:
        if os.path.exists(arquivo):
            caminho_historico = arquivo
            break
    
    if not caminho_historico:
        print("⚠️  Nenhuma planilha de histórico encontrada.")
        print("   Procurados:", ", ".join(arquivos_historico))
        usar_historico = input("   Deseja informar o caminho manualmente? (s/n): ").lower() == 's'
        
        if usar_historico:
            caminho_historico = input("   Digite o caminho da planilha: ").strip()
            if not os.path.exists(caminho_historico):
                print("   ❌ Arquivo não encontrado. Continuando sem histórico...")
                caminho_historico = None
    else:
        print(f"✅ Planilha de histórico encontrada: {caminho_historico}")
    
    analyzer = PortariaAnalyzer(caminho_historico)
    
    print("\n2. Escolhendo método de análise...")
    print("1. Analisar por URL")
    print("2. Analisar texto direto")
    
    opcao = input("Escolha uma opção (1 ou 2): ").strip()
    
    if opcao == "1":
        print("\n3. Obtendo URL da portaria...")
        url_portaria = input("Digite a URL da portaria para análise: ").strip()
        
        if not url_portaria:
            print("❌ URL não fornecida!")
            return
        
        print(f"URL informada: {url_portaria}")
        
        print("\n4. Iniciando análise...")
        print("-" * 40)
        
        # Analisar por URL
        try:
            resultado = analyzer.analisar_portaria(url_portaria)
            
            if 'erro' in resultado:
                print(f"❌ Erro: {resultado['erro']}")
            else:
                print(f"\n🎉 Análise concluída!")
                print(f"📊 {resultado['total_erros']} erros encontrados em {resultado['total_portarias']} portarias.")
                
                if resultado['arquivos_excel']:
                    print(f"📄 Relatórios Excel: {', '.join(resultado['arquivos_excel'])}")
                
                # Coletar todos os erros de todas as portarias
                todos_erros = []
                todas_pessoas = []
                
                for idx, res in enumerate(resultado['resultados'], 1):
                    if res['dados_portaria']:
                        dados = res['dados_portaria']
                        todas_pessoas.extend(dados['pessoas'])
                        
                        # Adicionar informações da portaria aos erros
                        for erro in res['erros']:
                            erro_completo = erro.copy()
                            erro_completo['portaria'] = dados['numero']
                            erro_completo['tipo_portaria'] = dados['tipo']
                            todos_erros.append(erro_completo)
                
                # Mostrar resumo consolidado
                print(f"\n{'='*80}")
                print("📋 RESUMO CONSOLIDADO DE TODOS OS ERROS ENCONTRADOS")
                print(f"{'='*80}")
                
                if todos_erros:
                    print(f"\n🔴 Total de {len(todos_erros)} erros encontrados:")
                    print("-" * 80)
                    
                    for i, erro in enumerate(todos_erros, 1):
                        pessoa = erro.get('pessoa', 'N/A')
                        tipo_erro = erro.get('tipo', 'ERRO DESCONHECIDO')
                        descricao = erro.get('descrição', '')
                        portaria = erro.get('portaria', 'N/A')
                        tipo_portaria = erro.get('tipo_portaria', 'N/A')
                        idade = erro.get('idade', '')
                        pais = erro.get('pais', '')
                        
                        print(f"{i:2d}. PESSOA: {pessoa}")
                        print(f"    📄 Portaria: {portaria} ({tipo_portaria})")
                        print(f"    ⚠️  Erro: {descricao}")
                        if idade:
                            print(f"    📅 Idade: {idade} anos")
                        if pais:
                            print(f"    🌍 País: {pais}")
                        print()
                else:
                    print("\n✅ Nenhum erro encontrado em nenhuma portaria!")
                
                # Mostrar estatísticas
                print(f"\n📊 ESTATÍSTICAS:")
                print(f"   • Total de portarias analisadas: {resultado['total_portarias']}")
                print(f"   • Total de pessoas processadas: {len(todas_pessoas)}")
                print(f"   • Total de erros encontrados: {len(todos_erros)}")
                
                # Contar tipos de erro
                tipos_erro = {}
                for erro in todos_erros:
                    tipo = erro.get('tipo', 'DESCONHECIDO')
                    tipos_erro[tipo] = tipos_erro.get(tipo, 0) + 1
                
                if tipos_erro:
                    print(f"\n📈 Tipos de erro encontrados:")
                    for tipo, quantidade in tipos_erro.items():
                        print(f"   • {tipo}: {quantidade}")
                
                print(f"\n{'='*80}")
        
        except KeyboardInterrupt:
            print("\n❌ Análise interrompida pelo usuário.")
        except Exception as e:
            print(f"\n❌ Erro inesperado: {e}")
            import traceback
            traceback.print_exc()
    
    elif opcao == "2":
        print("\n3. Inserindo texto da portaria...")
        print("Cole o texto da portaria abaixo (pressione Ctrl+D ou Ctrl+Z quando terminar):")
        
        linhas = []
        try:
            while True:
                linha = input()
                linhas.append(linha)
        except (EOFError, KeyboardInterrupt):
            pass
        
        texto_portaria = '\n'.join(linhas)
        
        if not texto_portaria.strip():
            print("❌ Texto não fornecido!")
            return
        
        print(f"\nTexto recebido: {len(texto_portaria)} caracteres")
        
        print("\n4. Iniciando análise...")
        print("-" * 40)
        
        # Analisar múltiplas portarias
        try:
            resultados, arquivos_excel = analyzer.analisar_multiplas_portarias(texto_portaria)
            
            # Coletar todos os erros de todas as portarias
            todos_erros = []
            todas_pessoas = []
            total_portarias = 0
            
            for idx, resultado in enumerate(resultados, 1):
                if resultado['dados_portaria']:
                    dados = resultado['dados_portaria']
                    todas_pessoas.extend(dados['pessoas'])
                    total_portarias += 1
                    
                    # Adicionar informações da portaria aos erros
                    for erro in resultado['erros']:
                        erro_completo = erro.copy()
                        erro_completo['portaria'] = dados['numero']
                        erro_completo['tipo_portaria'] = dados['tipo']
                        todos_erros.append(erro_completo)
            
            # Mostrar resumo consolidado
            print(f"\n{'='*80}")
            print("📋 RESUMO CONSOLIDADO DE TODOS OS ERROS ENCONTRADOS")
            print(f"{'='*80}")
            
            if todos_erros:
                print(f"\n🔴 Total de {len(todos_erros)} erros encontrados:")
                print("-" * 80)
                
                for i, erro in enumerate(todos_erros, 1):
                    pessoa = erro.get('pessoa', 'N/A')
                    tipo_erro = erro.get('tipo', 'ERRO DESCONHECIDO')
                    descricao = erro.get('descrição', '')
                    portaria = erro.get('portaria', 'N/A')
                    tipo_portaria = erro.get('tipo_portaria', 'N/A')
                    idade = erro.get('idade', '')
                    pais = erro.get('pais', '')
                    
                    print(f"{i:2d}. PESSOA: {pessoa}")
                    print(f"    📄 Portaria: {portaria} ({tipo_portaria})")
                    print(f"    ⚠️  Erro: {descricao}")
                    if idade:
                        print(f"    📅 Idade: {idade} anos")
                    if pais:
                        print(f"    🌍 País: {pais}")
                    print()
            else:
                print("\n✅ Nenhum erro encontrado em nenhuma portaria!")
            
            # Mostrar estatísticas
            print(f"\n📊 ESTATÍSTICAS:")
            print(f"   • Total de portarias analisadas: {total_portarias}")
            print(f"   • Total de pessoas processadas: {len(todas_pessoas)}")
            print(f"   • Total de erros encontrados: {len(todos_erros)}")
            
            # Contar tipos de erro
            tipos_erro = {}
            for erro in todos_erros:
                tipo = erro.get('tipo', 'DESCONHECIDO')
                tipos_erro[tipo] = tipos_erro.get(tipo, 0) + 1
            
            if tipos_erro:
                print(f"\n📈 Tipos de erro encontrados:")
                for tipo, quantidade in tipos_erro.items():
                    print(f"   • {tipo}: {quantidade}")
            
            if arquivos_excel:
                print(f"\n📄 Relatórios Excel gerados: {', '.join(arquivos_excel)}")
            
            print(f"\n{'='*80}")
        except KeyboardInterrupt:
            print("\n❌ Análise interrompida pelo usuário.")
        except Exception as e:
            print(f"\n❌ Erro inesperado: {e}")
            import traceback
            traceback.print_exc()
    
    else:
        print("❌ Opção inválida!")
        return
    
    print("\n" + "="*60)
    input("Pressione Enter para finalizar...")

if __name__ == "__main__":
    main()